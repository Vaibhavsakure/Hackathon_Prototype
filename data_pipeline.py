import pandas as pd
import numpy as np
import openpyxl
import os
import json

# ─── CONFIG ───────────────────────────────────────────────
PROCESS_FILE   = "data/_h_batch_process_data.xlsx"
PRODUCTION_FILE = "data/_h_batch_production_data.xlsx"
OUTPUT_CSV     = "data/master_df.csv"
INDIA_EMISSION_FACTOR = 0.82  # kg CO2 per kWh

# ─── STEP 1: Load production data ─────────────────────────
def load_production_data():
    df = pd.read_excel(PRODUCTION_FILE, sheet_name="BatchData")
    print(f"✅ Production data loaded: {len(df)} batches")
    return df

# ─── STEP 2: Load & aggregate process time-series data ────
def load_process_data():
    wb = openpyxl.load_workbook(PROCESS_FILE)
    batch_sheets = [s for s in wb.sheetnames if s.startswith("Batch_")]
    
    records = []
    for sheet_name in batch_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        data = pd.DataFrame(rows[1:], columns=headers)
        data = data.dropna(subset=["Batch_ID"])
        
        batch_id = data["Batch_ID"].iloc[0]
        
        # ── Total Energy ──
        total_energy = data["Power_Consumption_kW"].sum() / 60  # kWh (1 reading per minute)
        
        # ── Per-phase energy breakdown ──
        phase_energy = {}
        for phase, grp in data.groupby("Phase"):
            phase_energy[f"Energy_{phase}_kWh"] = grp["Power_Consumption_kW"].sum() / 60
        
        # ── Aggregate stats ──
        record = {
            "Batch_ID": batch_id,
            "Total_Energy_kWh": round(total_energy, 3),
            "Avg_Power_kW":     round(data["Power_Consumption_kW"].mean(), 3),
            "Max_Power_kW":     round(data["Power_Consumption_kW"].max(), 3),
            "Avg_Temperature":  round(data["Temperature_C"].mean(), 3),
            "Max_Temperature":  round(data["Temperature_C"].max(), 3),
            "Avg_Vibration":    round(data["Vibration_mm_s"].mean(), 4),
            "Max_Vibration":    round(data["Vibration_mm_s"].max(), 4),
            "Avg_Motor_Speed":  round(data["Motor_Speed_RPM"].mean(), 3),
            "Batch_Duration_min": len(data),
        }
        record.update(phase_energy)
        records.append(record)
    
    process_df = pd.DataFrame(records)
    print(f"✅ Process time-series aggregated: {len(process_df)} batches")
    return process_df

# ─── STEP 3: Merge & Feature Engineering ──────────────────
def build_master_df(production_df, process_df):
    master = pd.merge(production_df, process_df, on="Batch_ID", how="inner")
    
    # Carbon footprint
    master["Carbon_kg_CO2"] = (master["Total_Energy_kWh"] * INDIA_EMISSION_FACTOR).round(3)
    
    # ── Quality Score (0–100) ──
    # Higher Hardness = better, Higher Dissolution = better
    # Lower Friability = better, Higher Content_Uniformity = better (closer to 100)
    master["Quality_Score"] = (
        (master["Hardness"] / master["Hardness"].max() * 30) +
        (master["Dissolution_Rate"] / 100 * 30) +
        ((1 - master["Friability"] / master["Friability"].max()) * 20) +
        ((1 - abs(master["Content_Uniformity"] - 100) / 10) * 20)
    ).round(3)
    
    # ── Yield Score (0–100) ──
    # Lower Moisture = better, lower Friability = better
    master["Yield_Score"] = (
        ((1 - master["Moisture_Content"] / master["Moisture_Content"].max()) * 50) +
        ((1 - master["Friability"] / master["Friability"].max()) * 50)
    ).round(3)
    
    # ── Energy Efficiency Score ──
    # Higher quality per kWh = better
    master["Energy_Efficiency"] = (
        master["Quality_Score"] / master["Total_Energy_kWh"]
    ).round(4)
    
    # ── Performance Score (composite) ──
    master["Performance_Score"] = (
        master["Quality_Score"] * 0.4 +
        master["Yield_Score"] * 0.4 +
        master["Energy_Efficiency"] * 20
    ).round(3)
    
    print(f"✅ Master DataFrame built: {master.shape[0]} rows × {master.shape[1]} columns")
    print(f"\n📊 Key Stats:")
    print(f"   Avg Energy per batch : {master['Total_Energy_kWh'].mean():.2f} kWh")
    print(f"   Avg Carbon per batch : {master['Carbon_kg_CO2'].mean():.2f} kg CO₂")
    print(f"   Avg Quality Score    : {master['Quality_Score'].mean():.2f}/100")
    print(f"   Avg Yield Score      : {master['Yield_Score'].mean():.2f}/100")
    print(f"   Best batch (Quality) : {master.loc[master['Quality_Score'].idxmax(), 'Batch_ID']}")
    print(f"   Best batch (Energy)  : {master.loc[master['Total_Energy_kWh'].idxmin(), 'Batch_ID']}")
    
    return master

# ─── STEP 4: Save ─────────────────────────────────────────
def save_master(master):
    os.makedirs("data", exist_ok=True)
    master.to_csv(OUTPUT_CSV, index=False)
    print(f"\n✅ Saved → {OUTPUT_CSV}")
    print(f"   Columns: {list(master.columns)}")

# ─── MAIN ─────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 50)
    print("   AuraOptima — Data Pipeline")
    print("=" * 50)
    
    prod_df    = load_production_data()
    process_df = load_process_data()
    master     = build_master_df(prod_df, process_df)
    save_master(master)
    
    print("\n🎉 Pipeline complete! master_df.csv is ready.")